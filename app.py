import os
import json
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.image import Image
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.gridlayout import GridLayout
from kivy.clock import Clock
from datetime import datetime
from openpyxl import Workbook, load_workbook
from kivy.core.window import Window
from kivy.graphics.texture import Texture
from PIL import Image as PILImage
from kivy.uix.popup import Popup
import matplotlib.pyplot as plt





USER_DATA_FILE = "user_data.json"

class WelcomeScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        # Create the RelativeLayout to allow background image layering
        layout = RelativeLayout()

        # Load the background image (adjust file path as needed)
        try:
            background_image = Image(source='D:/python/fi.png', allow_stretch=True, keep_ratio=False)
            layout.add_widget(background_image)
        except Exception as e:
            print(f"Error loading image: {e}")

        # Add BoxLayout for widgets (on top of background image)
        box_layout = BoxLayout(orientation='vertical', padding=10, spacing=10, size_hint=(1, 0.5), pos_hint={'center_x': 0.5, 'center_y': 0.5})

        # Welcome label
        welcome_label = Label(text="Welcome to Fitness Tracking App", font_size=30, size_hint=(1, 0.5),)
        box_layout.add_widget(welcome_label)

        # Start button
        start_button = Button(text="Start", font_size=20, size_hint=(1, 0.5),background_color=(0.5, 0.5, 0.5, 0.5))
        start_button.bind(on_press=self.go_to_login)
        box_layout.add_widget(start_button)

        # Add the BoxLayout (widgets) on top of the image
        layout.add_widget(box_layout)

        # Add the entire layout to the screen
        self.add_widget(layout)

    def go_to_login(self, instance):
        self.manager.current = 'login'



class LoginScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Use RelativeLayout for layering the background image and widgets
        layout = RelativeLayout()

        # Add BoxLayout for input fields and buttons (positioned on top of the image)
        form_layout = BoxLayout(orientation='vertical', padding=10, spacing=10, size_hint=(0.8, 0.6), pos_hint={'center_x': 0.5, 'center_y': 0.5})

        # Check if user is already signed in
        self.message_label = Label(font_size=20, color=(1, 1, 1, 1))
        form_layout.add_widget(self.message_label)

        self.username_input = TextInput(hint_text="Username", font_size=20, size_hint=(1, 0.2), background_color=(0.7, 0.7, 0.7, 0.7))
        self.password_input = TextInput(hint_text="Password", font_size=20, size_hint=(1, 0.2), password=True, background_color=(0.7, 0.7, 0.7, 0.7))

        # Buttons for Login and Sign In
        self.login_button = Button(text="Login", font_size=20, size_hint=(1, 0.2), background_color=(0.3, 0.6, 1, 1))
        self.login_button.bind(on_press=self.login)

        self.sign_in_button = Button(text="Sign In", font_size=20, size_hint=(1, 0.2), background_color=(0.3, 0.8, 0.3, 1))
        self.sign_in_button.bind(on_press=self.sign_in)

        form_layout.add_widget(self.username_input)
        form_layout.add_widget(self.password_input)
        form_layout.add_widget(self.login_button)
        form_layout.add_widget(self.sign_in_button)

        # Add the form layout on top of the background
        layout.add_widget(form_layout)
        self.add_widget(layout)

        # Load user data
        self.user_data = self.load_user_data()
        self.check_signed_in()

    def load_user_data(self):
        """Load user data from a file."""
        if os.path.exists(USER_DATA_FILE):
            with open(USER_DATA_FILE, 'r') as f:
                return json.load(f)
        return {}

    def save_user_data(self):
        """Save user data to a file."""
        with open(USER_DATA_FILE, 'w') as f:
            json.dump(self.user_data, f)

    def check_signed_in(self):
        """Check if a user is already signed in."""
        if self.user_data.get("signed_in"):
            self.message_label.text = f"Welcome back, {self.user_data['signed_in']}! You're already signed in."
        else:
            self.message_label.text = "Please Sign In or Login."

    def login(self, instance):
        """Handle user login."""
        username = self.username_input.text.strip()
        password = self.password_input.text.strip()

        if not username or not password:
            self.message_label.text = "Please enter both username and password."
            return

        # Check if username exists and password matches
        if username in self.user_data and self.user_data[username] == password:
            self.user_data["signed_in"] = username
            self.save_user_data()
            self.message_label.text = f"Welcome back, {username}!"
            self.manager.current = 'bmi'
        else:
            self.message_label.text = "Invalid credentials. Please try again or sign in."

    def sign_in(self, instance):
        """Handle user sign-in."""
        username = self.username_input.text.strip()
        password = self.password_input.text.strip()

        if not username or not password:
            self.message_label.text = "Please enter both username and password."
            return

        # Check if username already exists
        if username in self.user_data:
            self.message_label.text = "Username already exists. Please login."
            return

        # Add new user to the data
        self.user_data[username] = password
        self.user_data["signed_in"] = username
        self.save_user_data()
        self.message_label.text = f"Welcome, {username}! You are now signed in."
        self.manager.current = 'bmi'
class BMIScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        self.weight_input = TextInput(hint_text="Weight (kg)", font_size=20, background_color=(0.5, 0.5, 0.5, 0.5))
        layout.add_widget(self.weight_input)

        self.height_input = TextInput(hint_text="Height (m)", font_size=20, background_color=(0.5, 0.5, 0.5, 0.5))
        layout.add_widget(self.height_input)

        self.calculate_button = Button(text="Calculate BMI", font_size=20, background_color=(0.3, 0.6, 1, 1))
        self.calculate_button.bind(on_press=self.calculate_bmi)
        layout.add_widget(self.calculate_button)

        self.result_label = Label(text="", font_size=20, color=(0.7, 0.7, 0.7, 1))
        layout.add_widget(self.result_label)

        self.next_button = Button(text="Next", font_size=20, background_color=(0.3, 0.6, 1, 1))
        self.next_button.bind(on_press=self.go_to_exercise)  # Bind the new go_to_exercise method
        layout.add_widget(self.next_button)

        self.back_button = Button(text="Back", font_size=20, background_color=(1, 0.3, 0.3, 1))
        self.back_button.bind(on_press=self.go_back)
        layout.add_widget(self.back_button)

        self.add_widget(layout)

        # New attribute to store BMI category
        self.result_category = None

    def calculate_bmi(self, instance):
        try:
            weight = float(self.weight_input.text)
            height = float(self.height_input.text)
            bmi = weight / (height ** 2)

            if bmi < 18.5:
                self.result_category = "Underweight"
                category = "Underweight"
                description = "You need to gain weight."
                tip = "Eat more calorie-dense foods like nuts and avocado."

            elif 18.5 <= bmi < 24.9:
                self.result_category = "Normal weight"
                category = "Normal weight"
                description = "You are in a healthy weight range."
                tip = "Maintain your weight with a balanced diet and regular exercise."

            elif 25 <= bmi < 29.9:
                self.result_category = "Overweight"
                category = "Overweight"
                description = "You need to lose some weight."
                tip = "Incorporate more physical activities like walking or jogging."

            else:
                self.result_category = "Obese"
                category = "Obese"
                description = "You need to lose weight to avoid health issues."
                tip = "Consult with a nutritionist for a tailored diet plan."

            # Display the result
            self.result_label.text = f"BMI: {bmi:.2f} ({category})\n{description}\nTip: {tip}"

        except ValueError:
            self.result_label.text = "Please enter valid numbers."

    # Add the go_to_exercise method
    def go_to_exercise(self, instance):
        # Get the ExerciseScreen instance and set its bmi_category attribute
        exercise_screen = self.manager.get_screen('exercise')
        exercise_screen.bmi_category = self.result_category
        self.manager.current = 'exercise'

    def go_back(self, instance):
        self.manager.current = 'login'  # Navigate back to login screen
class ExerciseScreen(Screen):
    def __init__(self, bmi_category=None, **kwargs):
        super().__init__(**kwargs)
        self.bmi_category = bmi_category  # Store the BMI category
        self.frames = []
        self.frame_index = 0  # Index to keep track of the current frame
        self.frame_duration = 0.1  # Duration for each frame (in seconds)

        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # Title label
        title_label = Label(text="Exercise Recommendations", font_size=30)
        layout.add_widget(title_label)

        # Example exercises (text)
        self.exercise_label = Label(
            text="Personalized exercise recommendations based on your BMI.",
            font_size=20
        )
        layout.add_widget(self.exercise_label)

        # Image widget to display the exercise GIF
        self.exercise_image = Image(allow_stretch=True)
        layout.add_widget(self.exercise_image)

        # Next button
        next_button = Button(text="Next", font_size=20, size_hint=(1, 0.5), background_color=(0.5, 0.5, 0.5, 0.5))
        next_button.bind(on_press=self.go_to_nutrition)
        layout.add_widget(next_button)

        # Back button
        back_button = Button(text="Back", font_size=20, size_hint=(1, 0.5), background_color=(1, 0.3, 0.3, 1))
        back_button.bind(on_press=self.go_back)
        layout.add_widget(back_button)

        self.add_widget(layout)

    def on_pre_enter(self, *args):
        # Set GIF size based on BMI category
        if self.bmi_category == "Underweight":
            self.exercise_image.size_hint = (1, 4)
            self.load_gif_frames('D:/python/sn.gif')
            self.exercise_label.text = "Focus on strength-building exercises."
            
        elif self.bmi_category == "Normal weight":
            self.exercise_image.size_hint = (1, 4)
            self.load_gif_frames('D:/python/Normalweight.gif')
            self.exercise_label.text = "Maintain a balanced exercise routine."
            
        elif self.bmi_category == "Overweight":
            self.exercise_image.size_hint = (1, 4)
            self.load_gif_frames('D:/python/overweight.gif')
            self.exercise_label.text = "Incorporate cardio and low-impact exercises."
            
        elif self.bmi_category == "Obese":
            self.exercise_image.size_hint = (1, 4)
            self.load_gif_frames('D:/python/obese.gif')
            self.exercise_label.text = "Start with low-intensity, joint-friendly exercises."
            
        else:
            self.exercise_image.source = ''
            self.exercise_label.text = "Exercise recommendations based on your BMI."

    def load_gif_frames(self, gif_path):
        # Open GIF using PIL and extract frames
        self.frames = []
        self.frame_index = 0
        try:
            pil_gif = PILImage.open(gif_path)
            while True:
                frame = pil_gif.convert("RGBA")  # Convert each frame to RGBA for Kivy compatibility
                texture = Texture.create(size=frame.size, colorfmt="rgba")
                texture.blit_buffer(frame.tobytes(), colorfmt="rgba", bufferfmt="ubyte")
                texture.flip_vertical()
                self.frames.append(texture)
                pil_gif.seek(pil_gif.tell() + 1)
        except EOFError:
            pass  # Reached the end of the GIF frames

        # Schedule frame updates
        if self.frames:
            Clock.schedule_interval(self.update_frame, self.frame_duration)

    def update_frame(self, dt):
        if self.frames:
            self.exercise_image.texture = self.frames[self.frame_index]
            self.frame_index = (self.frame_index + 1) % len(self.frames)

    def go_to_nutrition(self, instance):
        # Go to the Nutrition screen
        self.manager.current = 'nutrition'

    def go_back(self, instance):
        # Go back to BMI screen
        self.manager.current = 'bmi'

class NutritionScreen(Screen):
    def __init__(self, bmi, **kwargs):
        super().__init__(**kwargs)
        self.bmi = bmi
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # Title label
        title_label = Label(text="Nutrition Recommendations", font_size=30)
        layout.add_widget(title_label)

        # Nutrition information and images based on BMI
        if self.bmi < 18.5:  # Underweight
            nutrition_text = (
                "You may benefit from a diet higher in calories.\n"
                "Focus on healthy proteins and complex carbs."
            )
            img_source = "D:/python/Underweight.png"  # Update path as needed
        elif 18.5 <= self.bmi < 24.9:  # Normal weight
            nutrition_text = (
                "Continue with a balanced diet rich in fruits, vegetables, proteins, and carbs."
            )
            img_source = "D:/python/normalweight.png"  # Update path as needed
        elif 25 <= self.bmi < 29.9:  # Overweight
            nutrition_text = (
                "Consider reducing calorie intake while prioritizing\n"
                "nutrient-dense proteins and fibers."
            )
            img_source = "D:/python/overweight.png"  # Update path as needed
        else:  # Obese
            nutrition_text = (
                "Focus on portion control and low-calorie, high-protein foods.\n"
                "Incorporate more fiber to stay full longer."
            )
            img_source = "D:/python/obesity.png"  # Update path as needed

        # Debugging print statement
        print(f"BMI: {self.bmi}, Image Source: {img_source}")

        # Add image based on BMI
        try:
            bmi_image = Image(source=img_source, size_hint=(1, 4))
            layout.add_widget(bmi_image)
        except Exception as e:
            print(f"Error loading image: {e}")

        # Nutrition label
        nutrition_label = Label(text=nutrition_text, font_size=20)
        layout.add_widget(nutrition_label)

        # Next and Back buttons
        next_button = Button(text="Next", font_size=20, size_hint=(1, 0.5), background_color=(0.5, 0.5, 0.5, 0.5))
        next_button.bind(on_press=self.go_to_health)
        layout.add_widget(next_button)

        back_button = Button(text="Back", font_size=20, size_hint=(1, 0.5), background_color=(1, 0.3, 0.3, 1))
        back_button.bind(on_press=self.go_back)
        layout.add_widget(back_button)

        self.add_widget(layout)

    def go_to_health(self, instance):
        self.manager.current = 'health'

    def go_back(self, instance):
        self.manager.current = 'exercise'
class HealthScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=5, spacing=5)

        self.health_issue_input = TextInput(hint_text="Describe your health issue", font_size=20)
        layout.add_widget(self.health_issue_input)

        self.get_advice_button = Button(text="Get Advice", font_size=20)
        self.get_advice_button.bind(on_press=self.provide_advice)
        layout.add_widget(self.get_advice_button)

        self.advice_label = Label(text="", font_size=20)
        layout.add_widget(self.advice_label)

        self.exercise_image = Image(source='', allow_stretch=True, size_hint=(1, 0.5))
        layout.add_widget(self.exercise_image)

        self.next_button = Button(text="Next", font_size=20, size_hint=(1, 0.5), background_color=(0.5, 0.5, 0.5, 0.5))
        self.next_button.bind(on_press=self.go_to_report)
        layout.add_widget(self.next_button)


        self.back_button = Button(text="Back", font_size=20, background_color=(1, 0.3, 0.3, 1))  # Back button
        self.back_button.bind(on_press=self.go_back)
        layout.add_widget(self.back_button)

        self.add_widget(layout)

    def provide_advice(self, instance):
        health_issue = self.health_issue_input.text.lower()
        advice = ""
        exercise_image = ""

        if "diabetes" in health_issue:
            advice = (
                "Diabetes Management:\n"
                "1. Monitor blood sugar levels regularly.\n"
                "2. Follow a low-sugar, balanced diet.\n"
                "3. Engage in regular physical activity.\n"
                "4. Stay hydrated and avoid sugary drinks.\n"
                "5. Consult with a healthcare provider for medication management."
            )
            exercise_image = 'D:/python/d.png'  # Example image file

        elif "hypertension" in health_issue:
            advice = (
                "Hypertension Management:\n"
                "1. Reduce salt intake.\n"
                "2. Eat a diet rich in fruits and vegetables.\n"
                "3. Exercise regularly to maintain a healthy weight.\n"
                "4. Manage stress through relaxation techniques.\n"
                "5. Limit alcohol and avoid smoking."
            )
            exercise_image = 'D:/python/h.png'  # Example image file

        elif "cholesterol" in health_issue:
            advice = (
                "High Cholesterol Management:\n"
                "1. Avoid saturated and trans fats.\n"
                "2. Eat more fiber-rich foods.\n"
                "3. Exercise regularly to improve cholesterol levels.\n"
                "4. Maintain a healthy weight.\n"
                "5. Consider medication if prescribed by your doctor."
            )
            exercise_image = 'D:/python/c.png'  # Example image file

        elif "obesity" in health_issue:
            advice = (
                "Obesity Management:\n"
                "1. Limit processed foods, added sugars, and high-calorie snacks.\n"
                "2. Avoid eating straight from packages to prevent overeating.\n"
                "3. Drink plenty of water throughout the day. Sometimes, thirst is mistaken for hunger.\n"
                "4. Limit sugary drinks and alcohol.\n"
                "5. Aim for at least 15 minutes of moderate aerobic activity (like brisk walking) each day."
            )
            exercise_image = 'D:/python/o.png'  # Example image file

        else:
            advice = "Consult with a healthcare provider for tailored advice."

        self.advice_label.text = advice
        self.exercise_image.source = exercise_image
        self.exercise_image.reload()


    def go_to_report(self, instance):
        self.manager.current = 'report'
    def go_back(self, instance):
        self.manager.current = 'nutrition'

class ReportScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)

        # Title
        title_label = Label(text="Daily Report", font_size=30)
        layout.add_widget(title_label)

        # Inputs for exercise name, calories, and exercise duration
        self.exercise_name = TextInput(hint_text="Exercise Name", font_size=20)
        layout.add_widget(self.exercise_name)

        self.calories_burnt = TextInput(hint_text="Calories Burnt", font_size=20)
        layout.add_widget(self.calories_burnt)

        self.exercise_minutes = TextInput(hint_text="Minutes Exercised", font_size=20)
        layout.add_widget(self.exercise_minutes)

        # Save and Back Buttons
        save_button = Button(text="Save Report", font_size=20, background_color=(0.3, 0.6, 1, 1))
        save_button.bind(on_press=self.save_report)
        layout.add_widget(save_button)

        self.next_button = Button(text="Next", font_size=20, size_hint=(1, 0.5), background_color=(0.5, 0.5, 0.5, 0.5))
        self.next_button.bind(on_press=self.go_to_login)
        layout.add_widget(self.next_button)

        back_button = Button(text="Back", font_size=20, background_color=(1, 0.3, 0.3, 1))
        back_button.bind(on_press=self.go_back)
        layout.add_widget(back_button)

        # Button to generate and display the graph
        graph_button = Button(text="Show Monthly Graph", font_size=20, background_color=(0.3, 0.7, 0.3, 1))
        graph_button.bind(on_press=self.generate_graph)
        layout.add_widget(graph_button)

        self.add_widget(layout)

    def save_report(self, instance):
        # Collect data from inputs
        exercise_name = self.exercise_name.text
        calories = self.calories_burnt.text
        minutes = self.exercise_minutes.text

        # Current date and time
        current_datetime = datetime.now().strftime("%Y-%m-%d   %H:%M:%S")

        # File path for the Excel sheet
        excel_path = "Monthly_Report.xlsx"

        # Check if the Excel file exists
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)  # Load existing workbook
            sheet = workbook.active  # Get the active worksheet
        else:
            workbook = Workbook()  # Create a new workbook
            sheet = workbook.active
            # Add header row
            sheet.append(["Date & Time", "Exercise Name", "Calories Burnt", "Minutes Exercised"])

        # Add data to the Excel sheet
        sheet.append([current_datetime, exercise_name, calories, minutes])

        # Save the workbook
        workbook.save(excel_path)

        print(f"Report Saved: Exercise: {exercise_name}, Calories: {calories}, Minutes: {minutes}")
        print(f"Report saved to {excel_path}")

    def generate_graph(self, instance):
        # File path for the Excel sheet
        excel_path = "Monthly_Report.xlsx"

        if not os.path.exists(excel_path):
            print("No data available to generate the graph.")
            return

        # Load the Excel file
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Extract data
        dates = []
        calories = []

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            dates.append(row[0])  # Date & Time
            calories.append(float(row[2]))  # Calories Burnt

        # Generate graph
        plt.figure(figsize=(10, 6))
        plt.plot(dates, calories, marker='o', linestyle='-', color='b', label='Calories Burnt')
        plt.xlabel("Date")
        plt.ylabel("Calories")
        plt.title("Calories Burnt Over Time (Monthly)")
        plt.xticks(rotation=45, ha='right', fontsize=8)
        plt.tight_layout()
        plt.legend()
        graph_path = "monthly_graph.png"
        plt.savefig(graph_path)
        plt.close()

        # Display graph in a popup
        popup_layout = BoxLayout(orientation='vertical')
        popup_layout.add_widget(Image(source=graph_path))
        close_button = Button(text="Close", size_hint=(1, 0.2))
        close_button.bind(on_press=lambda *args: popup.dismiss())
        popup_layout.add_widget(close_button)

        popup = Popup(title="Monthly Graph", content=popup_layout, size_hint=(0.8, 0.8))
        popup.open()

    def go_to_login(self, instance):
        self.manager.current = 'login'

    def go_back(self, instance):
        # Navigate back to Health Screen
        self.manager.current = 'health'



class FitnessApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(WelcomeScreen(name='Welcome'))
        sm.add_widget(LoginScreen(name='login'))
        sm.add_widget(BMIScreen(name='bmi'))
        sm.add_widget(ExerciseScreen(name='exercise'))  # Add ExerciseScreen after BMIScreen
        sm.add_widget(NutritionScreen(name='nutrition', bmi=27))  # Replace 22 with the actual BMI value
        sm.add_widget(HealthScreen(name='health'))
        sm.add_widget(ReportScreen(name='report'))  # Add the new Report Screen

        return sm

if __name__ == '__main__':
    FitnessApp().run()
